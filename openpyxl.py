## This is an rough example of how to use openpyxl packages.
## openpyxl packages are very suitable for doing changes based on excel, but not a good tool to import excel to python. If you want to import
## excel to python, use pandas.read.excel

### reference https://automatetheboringstuff.com/chapter12/ & https://openpyxl.readthedocs.io/en/default/

import os
from openpyxl import load_workbook

# Set working dir
os.chdir('C:\\Users\\Chandler\\Desktop\\citi-project\\data')

wb = load_workbook('FGCI.xlsx')
print wb.get_sheet_names()
# Get all the worksheet and put them into a dictionary
ws_dict = {}
for i in wb.get_sheet_names():
    ws_dict["ws" + i] = wb[i]

for row in range(2,sheet.max_row+1):
    date = ws_dict['ws2.5']['A' + str()]
    
